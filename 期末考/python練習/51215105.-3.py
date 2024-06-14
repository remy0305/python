from openpyxl import load_workbook
import win32com.client as win32
import os

wb = load_workbook('51215105-2.xlsx')
ws = wb.active

TotalTradingVolume = 0
TotalTransactionVolume = 0

for row in range(2, 33):
    cell_value = ws['H' + str(row)].value
    if cell_value is not None:
        TotalTradingVolume += cell_value

        transaction_volume = cell_value * ws['I' + str(row)].value
        TotalTransactionVolume += transaction_volume

# Check if TotalTradingVolume is not zero to avoid division by zero
if TotalTradingVolume != 0:
    average_price = TotalTransactionVolume / TotalTradingVolume
else:
    average_price = 0

row += 3
# Update the values in the worksheet
ws[f'D{row}'] = TotalTradingVolume
ws[f'D{row+1}'] = average_price
ws[f'C{row}'] = '總交易量'
ws[f'C{row+1}'] = '總平均價'
ws[f'D{row+2}'] = "陳慶恩"

# Save the workbook
wb.save('51215105-3.xlsx')

# Create an Excel application object
excel = win32.Dispatch('Excel.Application')
excel.Visible = False

# Open the Excel workbook
workbook = excel.Workbooks.Open(os.getcwd() + '/' + '51215105-3.xlsx')
sheet = workbook.Sheets(1)

# Set print gridlines
sheet.PageSetup.PrintGridlines = True

